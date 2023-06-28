import os
from conftest import path_resources

from openpyxl import load_workbook, workbook


# оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    wb = load_workbook(
        os.path.join(path_resources, 'file_example_XLSX_50.xlsx'))
    sheet = wb.active
    sheet_value = sheet.cell(row=11, column=8).value
    assert sheet_value == 5486

    sheet_column = sheet.max_column
    assert sheet_column == 8
